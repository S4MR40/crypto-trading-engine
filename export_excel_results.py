import asyncio
import numpy as np
import pandas as pd
import pandas_ta as ta
import ccxt.async_support as ccxt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SYMBOLS = ["BTC/USDT", "ETH/USDT", "SOL/USDT"]
LIMIT_15M = 1500
INITIAL_BALANCE = 10000.0
RISK_PER_TRADE = 0.02
TAKER_FEE = 0.0005  # 0.05% per trade side
SLIPPAGE = 0.0002   # 0.02% estimated slippage

async def fetch_multi_timeframe_data(symbol: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    exchange = ccxt.binance({'enableRateLimit': True})
    try:
        ohlcv_1h = await exchange.fetch_ohlcv(symbol, timeframe="1h", limit=500)
        df_1h = pd.DataFrame(ohlcv_1h, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
        df_1h['timestamp'] = pd.to_datetime(df_1h['timestamp'], unit='ms')

        ohlcv_15m = await exchange.fetch_ohlcv(symbol, timeframe="15m", limit=LIMIT_15M)
        df_15m = pd.DataFrame(ohlcv_15m, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
        df_15m['timestamp'] = pd.to_datetime(df_15m['timestamp'], unit='ms')

        return df_1h, df_15m
    finally:
        await exchange.close()

def compute_indicators(df_1h: pd.DataFrame, df_15m: pd.DataFrame) -> pd.DataFrame:
    df_1h['ema_200_1h'] = ta.ema(df_1h['close'], length=200)
    
    df_1h_sorted = df_1h[['timestamp', 'ema_200_1h']].sort_values('timestamp')
    df_15m_sorted = df_15m.sort_values('timestamp')
    
    df = pd.merge_asof(
        df_15m_sorted,
        df_1h_sorted,
        on='timestamp',
        direction='backward'
    )
    
    df['rsi'] = ta.rsi(df['close'], length=14)
    df['atr'] = ta.atr(df['high'], df['low'], df['close'], length=14)
    
    adx_df = ta.adx(df['high'], df['low'], df['close'], length=14)
    df['adx'] = adx_df.iloc[:, 0]
    
    macd = ta.macd(df['close'], fast=12, slow=26, signal=9)
    df['macd_hist'] = macd.iloc[:, 1]
    
    return df.dropna().reset_index(drop=True)

def run_detailed_backtest(df: pd.DataFrame, symbol: str, config: dict) -> dict:
    balance = INITIAL_BALANCE
    position = None
    trades = []
    equity_curve = []

    rsi_l = config['rsi_l']
    rsi_h = config['rsi_h']
    sl_m = config['sl']
    tp_m = config['tp']
    min_adx = config['adx']
    be_trigger_mult = config['be']
    trail_mult = config['trail']

    for i in range(1, len(df)):
        curr = df.iloc[i]
        prev = df.iloc[i - 1]

        t_stamp = curr['timestamp']
        price = curr['close']
        high = curr['high']
        low = curr['low']
        open_p = curr['open']
        
        ema_200_1h = curr['ema_200_1h']
        rsi = curr['rsi']
        atr = curr['atr']
        adx = curr['adx']
        macd_curr = curr['macd_hist']
        macd_prev = prev['macd_hist']

        # 1. Active Position Trailing & Exits
        if position is not None:
            side = position['side']
            sl = position['sl']
            tp = position['tp']
            entry = position['entry_price']
            size = position['size']
            entry_atr = position['entry_atr']
            be_activated = position['be_activated']

            if side == "LONG":
                unrealized_profit = high - entry
                if not be_activated and (unrealized_profit >= entry_atr * be_trigger_mult):
                    sl = max(sl, entry)
                    position['be_activated'] = True
                    position['sl'] = sl
                
                if position['be_activated']:
                    new_sl = high - (atr * trail_mult)
                    if new_sl > position['sl']:
                        position['sl'] = new_sl
                        sl = new_sl

            elif side == "SHORT":
                unrealized_profit = entry - low
                if not be_activated and (unrealized_profit >= entry_atr * be_trigger_mult):
                    sl = min(sl, entry)
                    position['be_activated'] = True
                    position['sl'] = sl

                if position['be_activated']:
                    new_sl = low + (atr * trail_mult)
                    if new_sl < position['sl']:
                        position['sl'] = new_sl
                        sl = new_sl

            exit_triggered = False
            exit_price = price

            if side == "LONG":
                if open_p >= entry:
                    if high >= tp:
                        exit_triggered, exit_price = True, tp
                    elif low <= sl:
                        exit_triggered, exit_price = True, sl
                else:
                    if low <= sl:
                        exit_triggered, exit_price = True, sl
                    elif high >= tp:
                        exit_triggered, exit_price = True, tp

            elif side == "SHORT":
                if open_p <= entry:
                    if low <= tp:
                        exit_triggered, exit_price = True, tp
                    elif high >= sl:
                        exit_triggered, exit_price = True, sl
                else:
                    if high >= sl:
                        exit_triggered, exit_price = True, sl
                    elif low <= tp:
                        exit_triggered, exit_price = True, tp

            if exit_triggered:
                raw_pnl = (exit_price - entry) * size if side == "LONG" else (entry - exit_price) * size
                fee_cost = (entry * size * TAKER_FEE) + (exit_price * size * TAKER_FEE)
                slippage_cost = (entry + exit_price) * size * SLIPPAGE
                net_pnl = raw_pnl - fee_cost - slippage_cost
                
                balance += net_pnl
                trades.append({
                    "Symbol": symbol,
                    "Config": config['name'],
                    "Entry Time": position['entry_time'],
                    "Exit Time": t_stamp,
                    "Side": side,
                    "Entry Price": round(entry, 4),
                    "Exit Price": round(exit_price, 4),
                    "Position Size": round(size, 4),
                    "Net PnL ($)": round(net_pnl, 2),
                    "Return (%)": round((net_pnl / (entry * size)) * 100, 2) if size > 0 else 0
                })
                position = None

        # 2. MTF Entries
        if position is None and adx >= min_adx:
            if (price > ema_200_1h) and (rsi <= rsi_l) and (macd_curr > macd_prev):
                sl = price - (atr * sl_m)
                tp = price + (atr * tp_m)
                risk_amt = balance * RISK_PER_TRADE
                price_risk = abs(price - sl)
                size = (risk_amt / price_risk) if price_risk > 0 else 0.0
                position = {
                    "side": "LONG",
                    "entry_price": price,
                    "entry_time": t_stamp,
                    "sl": sl,
                    "tp": tp,
                    "size": size,
                    "entry_atr": atr,
                    "be_activated": False
                }

            elif (price < ema_200_1h) and (rsi >= rsi_h) and (macd_curr < macd_prev):
                sl = price + (atr * sl_m)
                tp = price - (atr * tp_m)
                risk_amt = balance * RISK_PER_TRADE
                price_risk = abs(price - sl)
                size = (risk_amt / price_risk) if price_risk > 0 else 0.0
                position = {
                    "side": "SHORT",
                    "entry_price": price,
                    "entry_time": t_stamp,
                    "sl": sl,
                    "tp": tp,
                    "size": size,
                    "entry_atr": atr,
                    "be_activated": False
                }

        equity_curve.append({
            "timestamp": t_stamp,
            "Symbol": symbol,
            "Config": config['name'],
            "Balance": round(balance, 2)
        })

    trades_df = pd.DataFrame(trades)
    eq_df = pd.DataFrame(equity_curve)

    total_trades = len(trades)
    if total_trades > 0:
        pnls = trades_df["Net PnL ($)"]
        wins = pnls[pnls > 0]
        losses = pnls[pnls < 0]

        win_rate = (len(wins) / total_trades) * 100
        gross_profit = wins.sum() if len(wins) > 0 else 0.0
        gross_loss = abs(losses.sum()) if len(losses) > 0 else 0.0
        profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else np.nan

        avg_win = wins.mean() if len(wins) > 0 else 0.0
        avg_loss = abs(losses.mean()) if len(losses) > 0 else 0.0
        win_loss_ratio = (avg_win / avg_loss) if avg_loss > 0 else np.nan

        eq_df["Peak"] = eq_df["Balance"].cummax()
        eq_df["Drawdown"] = (eq_df["Balance"] - eq_df["Peak"]) / eq_df["Peak"]
        max_drawdown = eq_df["Drawdown"].min() * 100

        returns = eq_df["Balance"].pct_change().dropna()
        sharpe = (returns.mean() / returns.std() * np.sqrt(35040)) if returns.std() > 0 else 0.0
    else:
        win_rate = gross_profit = gross_loss = profit_factor = avg_win = avg_loss = win_loss_ratio = max_drawdown = sharpe = 0.0

    stats = {
        "Config": config['name'],
        "Symbol": symbol,
        "Initial Balance": INITIAL_BALANCE,
        "Final Balance": round(balance, 2),
        "Net PnL ($)": round(balance - INITIAL_BALANCE, 2),
        "Return (%)": round(((balance - INITIAL_BALANCE) / INITIAL_BALANCE) * 100, 2),
        "Total Trades": total_trades,
        "Win Rate (%)": round(win_rate, 2),
        "Profit Factor": round(profit_factor, 2) if not np.isnan(profit_factor) else np.nan,
        "Max Drawdown (%)": round(max_drawdown, 2),
        "Sharpe Ratio": round(sharpe, 2),
        "Average Win ($)": round(avg_win, 2),
        "Average Loss ($)": round(avg_loss, 2),
        "Win/Loss Ratio": round(win_loss_ratio, 2) if not np.isnan(win_loss_ratio) else np.nan
    }

    return {"stats": stats, "trades": trades_df, "equity": eq_df}

def format_excel_report(filename: str):
    wb = openpyxl.load_workbook(filename)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws.views.sheetView[0].showGridLines = True

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(filename)

async def main():
    print("🚀 Running Multi-Asset Backtest & Generating Formatted Excel Report...\n")

    dfs = {}
    for sym in SYMBOLS:
        df_1h, df_15m = await fetch_multi_timeframe_data(sym)
        dfs[sym] = compute_indicators(df_1h, df_15m)

    configs = [
        {"name": "Fixed Stop (No BE)", "rsi_l": 40, "rsi_h": 60, "sl": 1.5, "tp": 2.5, "adx": 20, "be": 999.0, "trail": 999.0},
        {"name": "Break-Even @ +1.0x ATR", "rsi_l": 40, "rsi_h": 60, "sl": 1.5, "tp": 2.5, "adx": 20, "be": 1.0, "trail": 999.0},
        {"name": "BE + Trailing SL (1.5x ATR)", "rsi_l": 40, "rsi_h": 60, "sl": 1.5, "tp": 3.0, "adx": 20, "be": 1.0, "trail": 1.5},
    ]

    all_stats = []
    all_trades = []
    all_equity = []

    for cfg in configs:
        for sym in SYMBOLS:
            res = run_detailed_backtest(dfs[sym], sym, cfg)
            all_stats.append(res['stats'])
            if not res['trades'].empty:
                all_trades.append(res['trades'])
            all_equity.append(res['equity'])

    stats_df = pd.DataFrame(all_stats)
    trades_df = pd.concat(all_trades, ignore_index=True) if all_trades else pd.DataFrame()
    equity_df = pd.concat(all_equity, ignore_index=True) if all_equity else pd.DataFrame()

    excel_filename = "backtest_results.xlsx"
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        stats_df.to_excel(writer, sheet_name="Performance Statistics", index=False)
        trades_df.to_excel(writer, sheet_name="Trade Log", index=False)
        
        if not equity_df.empty:
            eq_pivot = equity_df.pivot_table(index="timestamp", columns=["Config", "Symbol"], values="Balance")
            eq_pivot.to_excel(writer, sheet_name="Equity Curves")

    # Apply styling
    format_excel_report(excel_filename)
    print(f"✅ Professional Excel report exported to: {excel_filename}")

if __name__ == "__main__":
    asyncio.run(main())
